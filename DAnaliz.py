import os
import sys
import shutil
import subprocess
import hashlib
import threading
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter.ttk import Progressbar, Style
from datetime import datetime

# ================== AYARLAR ==================
OUTPUT_EXCEL = "medya_analiz.xlsx"
SORUNLU_KLASOR = "SorunluDosyalar"

VIDEO_EXT = (".mp4", ".mkv", ".avi", ".mov", ".dav", ".264", ".h265", ".ts")
IMAGE_EXT = (".jpg", ".jpeg", ".png", ".bmp", ".tiff", ".webp")
AUDIO_EXT = (".mp3", ".wav", ".aac", ".ogg")

selected_folder = ""
last_report_path = None

# ================== PyInstaller Uyumlu Yol ==================
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS  # EXE içi
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

FFPROBE_PATH = resource_path(os.path.join("ffmpeg", "ffprobe.exe"))

# ================== VIDEO ANALİZ ==================
def get_video_info(filepath):
    try:
        cmd = [
            FFPROBE_PATH, "-v", "error",
            "-select_streams", "v:0",
            "-show_entries", "stream=codec_name,width,height",
            "-show_entries", "format=duration",
            "-of", "default=noprint_wrappers=1:nokey=1",
            filepath
        ]
        out = subprocess.check_output(cmd,stderr=subprocess.DEVNULL,creationflags=subprocess.CREATE_NO_WINDOW).decode().splitlines()
        codec, w, h, d = out[:4]
        d = float(d)
        if d <= 0:
            return codec, f"{w}x{h}", 0, True
        return codec, f"{w}x{h}", round(d, 2), False
    except Exception:
        return "", "", "", True

# ================== HASH ==================
def get_hash(path):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()

# ================== KLASÖR SEÇ ==================
def klasor_sec():
    global selected_folder
    folder = filedialog.askdirectory()
    if folder:
        selected_folder = folder
        folder_label.config(text=selected_folder)

# ================== RAPORU AÇ ==================
def raporu_ac():
    if last_report_path and os.path.exists(last_report_path):
        os.startfile(last_report_path)
    else:
        messagebox.showwarning("Uyarı", "Henüz oluşturulmuş bir rapor yok.")

# ================== ANALİZ THREAD ==================
def analiz_et():
    threading.Thread(target=_analiz_worker, daemon=True).start()

def _analiz_worker():
    global last_report_path

    if not selected_folder:
        messagebox.showwarning("Uyarı", "Lütfen analiz edilecek bir klasör seçin.")
        return

    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    medya_rows = []
    medya_disi_rows = []
    sorunlu_rows = []
    duplicate_rows = []
    file_records = {}

    all_files = []
    for root, _, files in os.walk(selected_folder):
        if SORUNLU_KLASOR in root:
            continue
        for f in files:
            all_files.append(os.path.join(root, f))

    toplam = len(all_files)
    progress["maximum"] = toplam
    progress["value"] = 0
    status.set("Analiz başladı…")

    bozuk_var = False

    for index, path in enumerate(all_files, start=1):
        progress["value"] = index

        yuzde = int((index / toplam) * 100)
        status.set(f"{index} / {toplam} dosya analiz ediliyor  —  %{yuzde}")

        app.update_idletasks()

        file = os.path.basename(path)
        ext = file.lower()
        klasor = os.path.dirname(path)

        try:
            stat = os.stat(path)
        except:
            continue

        tur = "Diğer"
        codec = ""
        coz = ""
        sure = ""
        bozuk = stat.st_size == 0

        is_video = ext.endswith(VIDEO_EXT)
        is_image = ext.endswith(IMAGE_EXT)
        is_audio = ext.endswith(AUDIO_EXT)
        is_media = is_video or is_image or is_audio

        if is_video:
            tur = "Video"
            codec, coz, sure, bozuk = get_video_info(path)
        elif is_image:
            tur = "Görsel"
        elif is_audio:
            tur = "Ses"

        if duplicate_var.get():
            try:
                h = get_hash(path)
                file_records.setdefault(h, {
                    "dosya": file,
                    "boyut": round(stat.st_size / 1024 / 1024, 2),
                    "klasorler": []
                })
                file_records[h]["klasorler"].append(klasor)
            except:
                pass

        if bozuk:
            if not bozuk_var:
                sorunlu_dir = os.path.join(selected_folder, SORUNLU_KLASOR)
                os.makedirs(sorunlu_dir, exist_ok=True)
                bozuk_var = True

            base, ext2 = os.path.splitext(file)
            sayac = 0
            yeni_yol = os.path.join(sorunlu_dir, file)
            while os.path.exists(yeni_yol):
                sayac += 1
                yeni_yol = os.path.join(sorunlu_dir, f"{base}_{sayac}{ext2}")
            shutil.move(path, yeni_yol)

            sorunlu_rows.append({
                "Dosya Adı": file,
                "Tür": tur,
                "İlk Tespit Edilen Klasör": klasor,
                "Boyut (MB)": round(stat.st_size / 1024 / 1024, 2)
            })

        if is_media:
            medya_rows.append({
                "Dosya Adı": file,
                "Tür": tur,
                "Codec": codec,
                "Çözünürlük": coz,
                "Süre (sn)": sure,
                "Boyut (MB)": round(stat.st_size / 1024 / 1024, 2),
                "Bozuk mu?": "EVET" if bozuk else "HAYIR",
                "İlk Tespit Edilen Klasör": klasor
            })

    if duplicate_var.get():
        for h, info in file_records.items():
            if len(info["klasorler"]) > 1:
                for k in info["klasorler"]:
                    duplicate_rows.append({
                        "Dosya Adı": info["dosya"],
                        "Boyut (MB)": info["boyut"],
                        "Hash": h,
                        "Klasör": k
                    })

    excel_path = os.path.join(selected_folder, OUTPUT_EXCEL)
    last_report_path = excel_path

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        if medya_rows:
            pd.DataFrame(medya_rows).to_excel(writer, index=False, sheet_name="MedyaDosyalari")
        if sorunlu_rows:
            pd.DataFrame(sorunlu_rows).to_excel(writer, index=False, sheet_name="SorunluDosyalar")
        if duplicate_rows:
            pd.DataFrame(duplicate_rows).to_excel(writer, index=False, sheet_name="TekrarlananDosyalar")

    if duplicate_rows:
        wb = load_workbook(excel_path)
        ws = wb["TekrarlananDosyalar"]
        fill = PatternFill(start_color="FFFFC0CB", end_color="FFFFC0CB", fill_type="solid")
        for r in range(2, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                ws.cell(r, c).fill = fill
        wb.save(excel_path)

    status.set(f"Analiz tamamlandı ✅  {toplam} / {toplam} dosya  —  %100")
    messagebox.showinfo("Tamamlandı", "Analiz tamamlandı.\nExcel raporu oluşturuldu.")

# ================== UI ==================
app = tk.Tk()
app.title("Klasördeki Dosya & Medya Analiz Aracı")

try:
    app.iconbitmap("DAnaliz.ico")
except:
    pass

w, h = 920, 600
sw, sh = app.winfo_screenwidth(), app.winfo_screenheight()
app.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")
app.configure(bg="#1e1e1e")
app.resizable(False, False)

style = Style()
style.theme_use("clam")
style.configure("dark.Horizontal.TProgressbar",
    troughcolor="#2b2b2b", background="#007acc", thickness=18)

tk.Label(app, text="Klasördeki Dosya & Medya Analiz Aracı",
    font=("Segoe UI", 16, "bold"),
    fg="#d4d4d4", bg="#1e1e1e").pack(pady=10)

duplicate_var = tk.BooleanVar(value=True)

tk.Button(app, text="📁 Klasör Seç",
    command=klasor_sec, bg="#3a3a3a", fg="white", width=24).pack(pady=8)

folder_label = tk.Label(app, text="Henüz klasör seçilmedi",
    wraplength=850, fg="#4FC1FF", bg="#1e1e1e")
folder_label.pack()

tk.Checkbutton(app, text="Duplicate (aynı dosya) analizi yap",
    variable=duplicate_var, bg="#1e1e1e",
    fg="#d4d4d4", selectcolor="#1e1e1e").pack(pady=(0, 10))

tk.Button(app, text="▶ Analiz Et",
    command=analiz_et, bg="#007acc", fg="white",
    font=("Segoe UI", 10, "bold"), width=14).pack(pady=10)

tk.Button(app, text="📊 Raporu Aç",
    command=raporu_ac, bg="#444444", fg="white", width=14).pack(pady=5)

progress = Progressbar(app, orient="horizontal",
    length=680, style="dark.Horizontal.TProgressbar")
progress.pack(pady=10)

status = tk.StringVar(value="Hazır")
tk.Label(app, textvariable=status,
    font=("Segoe UI", 11),
    fg="#4FC1FF", bg="#1e1e1e").pack(pady=5)

tk.Button(app, text="Uygulamayı Kapat",
    command=app.destroy, bg="#333333", fg="white", width=20).pack(pady=6)

tk.Label(app,
    text="* Analiz raporu Excel dosyası olarak, seçilen klasöre kayıt edilir\n"
         "** Sorunlu dosya varsa, Sorunlu Dosyalar klasörüne aktarılır",
    fg="#9CDCFE", bg="#1e1e1e", font=("Segoe UI", 9)).pack(pady=(0, 8))

tk.Label(app, text="© Mahmut ACAR 2026",
    fg="#4FC1FF", bg="#1e1e1e",
    font=("Segoe UI", 12, "bold")).pack(side="bottom", pady=6)

app.mainloop()